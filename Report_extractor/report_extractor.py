"""
사고조사보고서 자동 컬럼 추출기 v2
────────────────────────────────────────────────────────
주요 개선사항:
  1. 43개 필드를 6개 배치로 분할 추출 → 소형 LLM 부담 대폭 감소
  2. qwen3 계열 <think> 태그 자동 제거 + /no_think 프리픽스 적용
  3. JSON 스키마에서 주석(// ...) 제거 → 파싱 오류 원인 해결
  4. ChatOllama num_ctx=8192 설정 → 컨텍스트 초과 방지
  5. 배치 실패 시 규칙 기반(regex) 자동 폴백
  6. 실시간 배치별 진행률 표시
"""

import streamlit as st
import os, json, tempfile, re, io
from datetime import datetime
from copy import deepcopy

import pymupdf4llm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from langchain_community.chat_models import ChatOllama
    from langchain.schema import HumanMessage, SystemMessage
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False

# ──────────────────────────────────────────────────────────────
# 1. 컬럼 정의 + 배치 분할
# ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("발생일자",        "이벤트 발생 날짜. YYYY-MM-DD 형식. 예: 2025-03-23"),
    ("발생시간",        "이벤트 발생 시간. HH:MM 형식. 예: 07:49"),
    ("등록기관",        "데이터를 등록·보고한 기관명. 예: 서울교통공사, KORAIL"),
    ("철도구분",        "철도 유형. 일반철도/도시철도/고속철도 중 하나"),
    ("노선",           "노선명. 예: 서울 2호선, 경부선"),
    ("이벤트대분류",    "사건 유형 최상위 분류. 사고/장애/고장 중 하나"),
    ("이벤트중분류",    "발생 대상 분류. 차량/신호/선로/전력/외부요인 등"),
    ("이벤트소분류",    "상세 유형. 예: 탈선, 신호무응답, 차량고장"),
    ("주원인",         "1차 원인 요약. 짧고 명확하게 10~20자"),

    ("근본원인그룹",    "근본 원인 대분류. 인적요인/기술적요인/환경적요인 중 하나"),
    ("근본원인유형",    "근본 원인 범주. 예: 운전취급, 열차차량설비, 신호설비"),
    ("근본원인상세",    "상세 원인 설명. 보고서의 분석/결론 기반으로 구체적으로"),
    ("운행영향유형",    "운행에 미친 영향. 운행중단/지연운행/서행운전 등"),
    ("지연여부",        "지연 발생 여부. 지연/무지연 중 하나"),
    ("지연원인",        "지연 주요 원인. 사고/차량/신호/시설 등"),
    ("지연원인상세",    "지연 상세 사유. 구체적으로"),
    ("지연열차수",      "지연된 열차 수. 숫자만. 모르면 null"),
    ("최대지연시간(분)", "최대 지연 시간(분 단위). 숫자만"),

    ("총피해인원",      "사망자+부상자 합계. 숫자만"),
    ("사망자수",        "사망자 수. 숫자만. 없으면 0"),
    ("부상자수",        "부상자 수. 숫자만. 없으면 0"),
    ("피해액(백만원)",  "재산 피해액 백만원 단위. 숫자만"),
    ("행정구역",        "사고 발생 행정 주소. 가능한 상세히"),
    ("발생역A",         "기준역 또는 주요역 이름"),
    ("발생역B",         "인접역 또는 상대역 이름. 없으면 발생역A와 동일"),
    ("장소대분류",      "장소 유형. 역/본선/기지/차량기지 등"),

    ("장소중분류",      "세부 장소. 구내선로/본선/승강장 등"),
    ("상세위치",        "추가 위치 설명. 예: 4번 승강장, 32km 지점"),
    ("기상상태",        "사건 당시 기상. 맑음/흐림/비/눈/안개 등"),
    ("온도",           "기온(℃). 숫자만"),
    ("강우량",         "강우량(mm). 숫자만. 없으면 0"),
    ("적설량",         "적설량(cm). 숫자만. 없으면 0"),
    ("대상구분",        "대상 구분. 열차/차량/설비 중 하나"),
    ("열차종류",        "열차 유형. 전동열차/화물열차/여객열차/KTX 등"),

    ("선로유형",        "선로 형태. 지상/지하/교량 중 하나"),
    ("신호시스템유형",  "신호 시스템 종류. 예: ATP/ATO, 자동폐색, ETCS"),
    ("고장부품명",      "문제 발생 부품명. 해당없으면 해당없음"),
    ("고장현상",        "고장·이상 현상 설명. 해당없으면 해당없음"),
    ("고장원인",        "기술적 고장 원인. 해당없으면 해당없음"),
    ("조치내용",        "현장·정비 조치 내용 요약"),
    ("이벤트개요",      "발생 일시·장소·영향을 포함한 요약. 3~5문장"),
    ("직접원인",        "사고·장애 직접 원인. 조사위원회 결론 문장 기반"),
    ("데이터 출처",     "보고서 문서명 또는 출처"),
]

# 6개 배치로 분할 (7~8필드씩)
BATCHES = [
    COLUMNS[0:9],    # 기본정보: 날짜~주원인
    COLUMNS[9:18],   # 원인·지연: 근본원인그룹~최대지연시간
    COLUMNS[18:26],  # 피해·위치: 총피해~장소대분류
    COLUMNS[26:34],  # 위치·기상·대상
    COLUMNS[34:43],  # 선로·신호·고장·조치·개요·원인
]
BATCH_NAMES = ["기본정보", "원인·지연", "피해·위치A", "위치·기상", "선로·고장·개요"]

COLUMN_NAMES = [c[0] for c in COLUMNS]


# ──────────────────────────────────────────────────────────────
# 2. LLM 유틸리티
# ──────────────────────────────────────────────────────────────
def is_qwen3(model_name: str) -> bool:
    return "qwen3" in model_name.lower()

def clean_llm_output(raw: str) -> str:
    """qwen3 <think> 태그 및 코드블록 제거"""
    # <think>...</think> 완전 제거 (qwen3 CoT)
    raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL)
    # 마크다운 코드블록 제거
    raw = re.sub(r"```(?:json)?", "", raw)
    raw = raw.replace("```", "")
    return raw.strip()

def safe_json_parse(text: str) -> dict:
    """다중 전략으로 JSON 파싱"""
    text = clean_llm_output(text)

    # 1차: 직접 파싱
    try:
        return json.loads(text)
    except Exception:
        pass

    # 2차: 첫 번째 { } 블록 추출
    match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)?\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except Exception:
            pass

    # 3차: 더 관대한 { ~ } 추출 (중첩 허용)
    start = text.find('{')
    end = text.rfind('}')
    if start != -1 and end != -1 and end > start:
        candidate = text[start:end+1]
        # JS 주석 제거
        candidate = re.sub(r'//[^\n]*', '', candidate)
        # trailing comma 제거
        candidate = re.sub(r',(\s*[}\]])', r'\1', candidate)
        try:
            return json.loads(candidate)
        except Exception:
            pass

    # 4차: 키-값 패턴 수동 파싱
    result = {}
    for key, _ in COLUMNS:
        pattern = rf'"{re.escape(key)}"\s*:\s*"([^"]*)"'
        m = re.search(pattern, text)
        if m:
            result[key] = m.group(1)
        else:
            pattern_num = rf'"{re.escape(key)}"\s*:\s*(-?\d+(?:\.\d+)?)'
            m2 = re.search(pattern_num, text)
            if m2:
                result[key] = m2.group(1)
    return result


def build_batch_prompt(batch_cols: list, report_text: str, model_name: str) -> str:
    prefix = "/no_think\n" if is_qwen3(model_name) else ""

    # 간결한 예시 포함 스키마 (주석 없이 순수 JSON 구조)
    schema_keys = ", ".join(f'"{name}"' for name, _ in batch_cols)
    field_guide = "\n".join(
        f'  - "{name}": {desc}'
        for name, desc in batch_cols
    )

    # 보고서 텍스트 - 배치당 5000자 (중복 없이)
    text_chunk = report_text[:6000]

    prompt = f"""{prefix}당신은 철도사고 보고서를 분석하는 전문가입니다.
아래 [보고서]에서 지정한 필드만 추출하여 JSON으로 출력하세요.

[추출 규칙]
- 반드시 JSON 객체만 출력하세요. 설명·사유·주석 절대 금지.
- 정보 없으면 null 사용 (빈 문자열 사용 금지).
- 날짜: YYYY-MM-DD, 시간: HH:MM, 숫자 필드: 숫자만.

[추출할 필드 ({len(batch_cols)}개)]
{field_guide}

[출력 형식]
{{{schema_keys}}}

[보고서]
{text_chunk}

JSON:"""
    return prompt


def call_llm_batch(llm, batch_cols: list, report_text: str, model_name: str) -> dict:
    system = SystemMessage(content=(
        "Output ONLY a valid JSON object. "
        "No thinking, no explanation, no markdown fences. Just JSON."
    ))
    prompt = build_batch_prompt(batch_cols, report_text, model_name)
    response = llm.invoke([system, HumanMessage(content=prompt)])
    return safe_json_parse(response.content)


# ──────────────────────────────────────────────────────────────
# 3. 규칙 기반 폴백 추출 (regex)
# ──────────────────────────────────────────────────────────────
def regex_extract(report_text: str) -> dict:
    data: dict = {}
    t = report_text

    # 날짜/시간
    dm = re.search(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', t)
    if dm:
        data['발생일자'] = f"{dm.group(1)}-{int(dm.group(2)):02d}-{int(dm.group(3)):02d}"
    tm = re.search(r'(\d{1,2})시\s*(\d{2})분경?', t)
    if tm:
        data['발생시간'] = f"{int(tm.group(1)):02d}:{tm.group(2)}"

    # 기관
    if '서울교통공사' in t: data['등록기관'] = '서울교통공사'
    elif 'KORAIL' in t or '한국철도' in t: data['등록기관'] = 'KORAIL'

    # 철도구분
    if any(k in t for k in ['지하철', '도시철도', '호선']): data['철도구분'] = '도시철도'
    elif '고속철도' in t or 'KTX' in t: data['철도구분'] = '고속철도'
    else: data['철도구분'] = '일반철도'

    # 노선
    nm = re.search(r'(서울\s*\d+호선|경부선|경부고속선|중앙선|경인선|수인선)', t)
    if nm: data['노선'] = nm.group(1).replace(' ', ' ')

    # 이벤트 분류
    if '탈선' in t: data['이벤트대분류']='사고'; data['이벤트중분류']='차량'; data['이벤트소분류']='탈선'
    elif '충돌' in t: data['이벤트대분류']='사고'; data['이벤트중분류']='차량'; data['이벤트소분류']='충돌'
    elif '화재' in t: data['이벤트대분류']='사고'; data['이벤트중분류']='차량'; data['이벤트소분류']='화재'

    # 피해
    dead = re.search(r'사망자?\s*수?\s*[：:]\s*(\d+)|인명피해는\s*없었다', t)
    if dead:
        data['사망자수'] = '0' if '없었' in dead.group(0) else dead.group(1)
    injured = re.search(r'부상자?\s*수?\s*[：:]\s*(\d+)', t)
    data['부상자수'] = injured.group(1) if injured else '0'
    data['총피해인원'] = str(int(data.get('사망자수','0')) + int(data.get('부상자수','0')))

    # 피해액
    dmg = re.search(r'총\s*([\d,]+)\s*백만\s*원|피해금액\s*([\d,]+)', t)
    if dmg:
        val = (dmg.group(1) or dmg.group(2)).replace(',', '')
        data['피해액(백만원)'] = val

    # 기상
    weather_map = {'맑았다': '맑음', '맑음': '맑음', '흐림': '흐림', '비': '비', '눈': '눈', '안개': '안개'}
    for kw, wv in weather_map.items():
        if kw in t: data['기상상태'] = wv; break

    # 온도
    temp = re.search(r'기온은?\s*([-\d.]+)\s*℃|온도\s*([-\d.]+)', t)
    if temp: data['온도'] = (temp.group(1) or temp.group(2))

    # 선로유형
    if '지하' in t: data['선로유형'] = '지하'
    elif '교량' in t: data['선로유형'] = '교량'
    else: data['선로유형'] = '지상'

    # 신호시스템
    if 'ATP' in t and 'ATO' in t: data['신호시스템유형'] = 'ATP/ATO'
    elif '자동폐색' in t: data['신호시스템유형'] = '자동폐색'

    return data


# ──────────────────────────────────────────────────────────────
# 4. 메인 추출 함수 (배치 + 폴백 결합)
# ──────────────────────────────────────────────────────────────
def extract_all(report_text: str, model_name: str, progress_fn=None) -> dict:
    """배치 LLM 추출 + regex 폴백 병합"""

    # regex 기반 기초값 먼저 채움
    result = regex_extract(report_text)

    if not LLM_AVAILABLE:
        return result

    llm = ChatOllama(
        model=model_name,
        base_url="http://127.0.0.1:11434",
        temperature=0,
        num_ctx=8192,          # 컨텍스트 윈도우 명시
        num_predict=2048,      # 출력 토큰 제한
    )

    total_batches = len(BATCHES)
    batch_errors = []

    for i, batch in enumerate(BATCHES):
        batch_name = BATCH_NAMES[i]
        pct = 0.15 + (0.65 * i / total_batches)

        if progress_fn:
            progress_fn(pct, f"🔍 배치 {i+1}/{total_batches}: {batch_name} 추출 중...")

        try:
            batch_result = call_llm_batch(llm, batch, report_text, model_name)

            # LLM 결과로 덮어쓰기 (null/빈값 제외)
            for col_name, _ in batch:
                val = batch_result.get(col_name)
                if val is not None and str(val).strip() not in ("", "null", "NULL", "None"):
                    result[col_name] = str(val).strip()

        except Exception as e:
            batch_errors.append(f"배치{i+1}({batch_name}): {e}")
            # 실패해도 계속 진행 (이미 regex 값이 있음)

    # 파생 필드 자동 계산
    try:
        dead = int(result.get('사망자수', 0) or 0)
        inj = int(result.get('부상자수', 0) or 0)
        result['총피해인원'] = str(dead + inj)
    except Exception:
        pass

    # 데이터 출처 기본값
    if not result.get('데이터 출처'):
        result['데이터 출처'] = 'PDF 보고서 자동 추출'

    if progress_fn:
        progress_fn(0.85, "📊 Excel 생성 중...")

    if batch_errors:
        result['_batch_errors'] = " | ".join(batch_errors)

    return result


# ──────────────────────────────────────────────────────────────
# 5. Excel 생성
# ──────────────────────────────────────────────────────────────
def create_excel(extracted: dict, source_name: str, event_type: str = "사고") -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill("solid", start_color="EEF4FB")
    null_fill   = PatternFill("solid", start_color="FFF3CD")  # 미추출 필드 강조

    # ── 결과 시트 ──
    ws = wb.create_sheet(event_type)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75
    ws.row_dimensions[1].height = 28

    for col, title in enumerate(["컬럼명", "추출값"], 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font; c.fill = header_fill
        c.alignment = center_al; c.border = border

    for r, (col_name, col_desc) in enumerate(COLUMNS, 2):
        val_str = extracted.get(col_name)
        val_str = "" if (val_str is None or str(val_str).strip() in ("null","NULL","None")) else str(val_str).strip()
        is_empty = val_str == ""
        ws.row_dimensions[r].height = 20 if len(val_str) < 60 else 35

        a = ws.cell(row=r, column=1, value=col_name)
        a.font = Font(bold=True, name="Arial", size=10, color="1F4E79")
        a.alignment = Alignment(horizontal="left", vertical="center")
        a.border = border
        if is_empty:
            a.fill = null_fill
        elif r % 2 == 0:
            a.fill = alt_fill

        b = ws.cell(row=r, column=2, value=val_str)
        b.font = Font(name="Arial", size=10, color="555555" if is_empty else "000000")
        b.alignment = left_wrap
        b.border = border
        if is_empty:
            b.fill = null_fill
            b.value = "⬜ 미추출"
        elif r % 2 == 0:
            b.fill = alt_fill

    # 메타 정보
    last = len(COLUMNS) + 3
    for offset, (label, val) in enumerate([
        ("추출 모델", extracted.get("_model", "-")),
        ("추출 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("원본 파일", source_name),
    ]):
        ws.cell(row=last+offset, column=1, value=label).font = Font(italic=True, color="888888", name="Arial", size=9)
        ws.cell(row=last+offset, column=2, value=val).font = Font(italic=True, color="888888", name="Arial", size=9)

    # ── 컬럼정의서 시트 ──
    ws2 = wb.create_sheet("컬럼정의서")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55
    for col, t in enumerate(["컬럼명", "설명"], 1):
        c = ws2.cell(row=1, column=col, value=t)
        c.font = header_font; c.fill = header_fill
        c.alignment = center_al; c.border = border
    for r, (n, d) in enumerate(COLUMNS, 2):
        ws2.cell(row=r, column=1, value=n).font = Font(bold=True, name="Arial", size=10, color="1F4E79")
        ws2.cell(row=r, column=2, value=d).font = Font(name="Arial", size=10)
        ws2.cell(row=r, column=1).border = border
        ws2.cell(row=r, column=2).border = border
        if r % 2 == 0:
            ws2.cell(row=r, column=1).fill = alt_fill
            ws2.cell(row=r, column=2).fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────────────────────
# 6. Streamlit UI
# ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="📄 보고서 자동 입력기 v2", layout="wide")
st.title("📄 사고조사보고서 → 컬럼 자동 추출기 v2")
st.caption("PDF 보고서를 업로드하면 LLM이 **5회 배치 추출** + regex 폴백으로 43개 컬럼을 채웁니다.")

# ── 사이드바 ──
with st.sidebar:
    st.header("⚙️ 설정")

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    CONFIG_FILE = os.path.join(BASE_DIR, "shared", "system_config.json")
    default_model = "qwen2.5:7b-instruct"
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                default_model = json.load(f).get("selected_model", default_model)
        except Exception:
            pass

    ollama_models = [
        "qwen3:8b",
        "qwen2.5:7b-instruct",
    ]
    try:
        idx = ollama_models.index(default_model)
    except ValueError:
        idx = 1

    model_name = st.selectbox("🤖 LLM 모델", ollama_models, index=idx)

    # qwen3 사용 시 안내
    if is_qwen3(model_name):
        st.info("💡 qwen3 감지: `/no_think` 자동 적용 및 `<think>` 태그 자동 제거")

    st.divider()
    event_type = st.selectbox("이벤트 유형 (시트명)", ["사고", "장애", "고장"])

    st.divider()
    with st.expander("🔧 고급 설정"):
        use_llm = st.checkbox("LLM 사용", value=True,
                              help="체크 해제 시 regex만 사용 (빠르지만 품질 낮음)")
        show_raw = st.checkbox("LLM 원문 응답 보기", value=False)

    st.divider()
    st.markdown("""
**배치 추출 전략**
1. 기본정보 (날짜~주원인)
2. 원인·지연
3. 피해·위치A
4. 위치·기상
5. 선로·고장·개요

각 배치가 실패해도 regex 폴백이 보완합니다.
""")

# ── 파일 업로드 ──
uploaded = st.file_uploader(
    "📂 PDF 업로드 (사고조사보고서)",
    type=["pdf"],
    help="항공·철도사고조사위원회 등의 공식 보고서 PDF"
)

if uploaded:
    c1, c2 = st.columns([4, 1])
    with c1:
        st.success(f"✅ **{uploaded.name}** ({uploaded.size / 1024:.1f} KB) 로드 완료")
    with c2:
        run_btn = st.button("🚀 추출 시작", type="primary", use_container_width=True)

    if run_btn:
        prog = st.progress(0.0)
        stat = st.empty()
        debug_area = st.expander("📋 배치별 추출 로그", expanded=False)
        log_lines = []

        def update(pct, msg):
            prog.progress(pct)
            stat.info(msg)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name

        try:
            update(0.05, "📖 PDF 텍스트 추출 중...")
            report_text = pymupdf4llm.to_markdown(tmp_path)
            st.session_state['report_text'] = report_text
            update(0.12, f"✅ 텍스트 추출 완료 ({len(report_text):,}자) — 배치 LLM 추출 시작...")

            # 배치 추출 실행
            if use_llm and LLM_AVAILABLE:
                extracted = extract_all(report_text, model_name, update)
            else:
                extracted = regex_extract(report_text)
                extracted['데이터 출처'] = uploaded.name

            extracted['_model'] = model_name
            excel_bytes = create_excel(extracted, uploaded.name, event_type)

            prog.progress(1.0)
            stat.success("🎉 추출 완료!")

            st.session_state.update({
                'extracted': extracted,
                'excel_bytes': excel_bytes,
                'source_name': uploaded.name,
            })
            st.rerun()

        except Exception as e:
            import traceback
            st.error(f"오류 발생: {e}")
            st.text(traceback.format_exc())
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

# ── 결과 출력 ──
if 'extracted' in st.session_state:
    extracted   = st.session_state['extracted']
    excel_bytes = st.session_state['excel_bytes']

    st.divider()

    # KPI
    total  = len(COLUMN_NAMES)
    filled = sum(1 for k in COLUMN_NAMES
                 if extracted.get(k) not in (None, "", "null", "NULL", "None", "⬜ 미추출"))
    rate   = filled / total * 100

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("전체 컬럼", f"{total}개")
    m2.metric("추출 성공", f"{filled}개")
    m3.metric("추출률", f"{rate:.0f}%")
    m4.metric("LLM 모델", extracted.get("_model", "-"))

    # 다운로드
    fname = f"사고보고서_추출_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        "⬇️ Excel 다운로드",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # 배치 오류 표시
    if '_batch_errors' in extracted:
        st.warning(f"⚠️ 일부 배치 오류 (regex로 보완됨): {extracted['_batch_errors']}")

    # 결과 테이블
    st.markdown("### 📋 추출 결과")
    import pandas as pd
    rows = []
    for col_name, col_desc in COLUMNS:
        val = extracted.get(col_name)
        val_str = "" if val in (None, "null", "NULL", "None") else str(val)
        ok = val_str and val_str != "⬜ 미추출"
        rows.append({
            "상태": "✅" if ok else "⬜",
            "컬럼명": col_name,
            "추출값": (val_str[:100] + "...") if len(val_str) > 100 else val_str,
            "설명": col_desc,
        })
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True, height=720,
                 column_config={
                     "상태": st.column_config.TextColumn("", width=40),
                     "컬럼명": st.column_config.TextColumn("컬럼명", width=130),
                     "추출값": st.column_config.TextColumn("추출값", width=400),
                     "설명": st.column_config.TextColumn("설명", width=220),
                 })

    # 원문
    if show_raw and 'report_text' in st.session_state:
        with st.expander("📄 원문 텍스트 (처음 4000자)"):
            st.text(st.session_state['report_text'][:4000])

else:
    st.info("👆 PDF 파일을 업로드하고 '추출 시작' 버튼을 클릭하세요.")
    with st.expander("📋 추출 대상 컬럼 목록 (43개)"):
        import pandas as pd
        st.dataframe(pd.DataFrame(COLUMNS, columns=["컬럼명","설명"]),
                     use_container_width=True, hide_index=True)