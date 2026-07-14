"""엑셀 원본 데이터를 PostgreSQL로 적재하는 ETL 스크립트.

사용법:
    python -m etl.load_data --pm25 /data/pm25_202410.xlsx --traffic /data/traffic_line1_202604.xlsx

기대하는 원본 포맷:
  - PM2.5 파일: '공단양식' 시트 (컬럼: 역명, 지점코드, 측정시간(YYYYMMDDHH), 항목코드, 시간, 측정값, ...)
  - 통행량 파일: '1호선' 시트 (역명 / 구분(승차·하차) / 계 / 시간대별 05시~24시이후 컬럼)

⚠ 통행량 원본 마지막 두 행('1호선 계')은 노선 전체 합계이며 개별 역 데이터가 아니므로 반드시 제외한다.
   (이전 버전에서 병합 셀 처리 로직 오류로 마지막 역인 '송도달빛축제공원'의 하차값이
    노선 전체 합계로 덮어써지는 버그가 있었음 — 아래 로직은 이를 수정한 버전)
"""
from __future__ import annotations
import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

sys.path.append(str(Path(__file__).resolve().parents[1]))
from app.database import engine, SessionLocal, Base
from app import models

# 인천1호선 33개 역 순서 및 근사 좌표 (인천교통공사 공식 역정보 도로명주소 기반 근사치).
# 정밀 GPS 좌표(공공데이터포털 CSV)는 브라우저 다운로드 상호작용이 필요해 자동 수집이 불가했음 — 참고용.
LINE1_COORDS: dict[str, tuple[float, float]] = {
    "검단호수공원": (37.6068, 126.6683), "신검단중앙": (37.6015, 126.6751), "아라": (37.5910, 126.6884),
    "계양": (37.5713, 126.7368), "귤현": (37.5804, 126.7460), "박촌": (37.5875, 126.7379),
    "임학": (37.5606, 126.7361), "계산": (37.5457, 126.7286), "경인교대입구": (37.5385, 126.7248),
    "작전": (37.5310, 126.7223), "갈산": (37.5218, 126.7238), "부평구청": (37.5089, 126.7218),
    "부평시장": (37.4949, 126.7239), "부평": (37.4897, 126.7241), "동수": (37.4820, 126.7291),
    "부평삼거리": (37.4778, 126.7326), "간석오거리": (37.4735, 126.7386), "인천시청": (37.4562, 126.7052),
    "예술회관": (37.4489, 126.7011), "인천터미널": (37.4426, 126.7002), "문학경기장": (37.4370, 126.6997),
    "선학": (37.4322, 126.6935), "신연수": (37.4204, 126.6822), "원인재": (37.4133, 126.6789),
    "동춘": (37.4067, 126.6716), "동막": (37.3961, 126.6591), "캠퍼스타운": (37.3855, 126.6510),
    "테크노파크": (37.3820, 126.6462), "지식정보단지": (37.3813, 126.6389), "인천대입구": (37.3752, 126.6327),
    "센트럴파크": (37.3931, 126.6379), "국제업무지구": (37.3930, 126.6300), "송도달빛축제공원": (37.3899, 126.6178),
}


def parse_pm25(xlsx_path: str) -> dict[str, list[tuple[dt.datetime, float]]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["공단양식"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # header skip
    out: dict[str, list[tuple[dt.datetime, float]]] = {}
    for r in rows:
        station, _code, mtime, _item, _hh, val = r[0], r[1], r[2], r[3], r[4], r[5]
        if station is None or mtime is None or val is None:
            continue
        mtime = str(mtime)
        ts = dt.datetime.strptime(mtime, "%Y%m%d%H")
        out.setdefault(station, []).append((ts, float(val)))
    wb.close()
    for st in out:
        out[st].sort(key=lambda t: t[0])
    return out


def parse_traffic(xlsx_path: str) -> tuple[list[str], dict[str, dict[str, list[int]]]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["1호선 "]
    rows = list(ws.iter_rows(values_only=True))
    hour_labels = [str(h) for h in rows[4][3:23]]

    out: dict[str, dict[str, list[int]]] = {}
    i = 5
    cur_station = None
    while i < len(rows):
        r = rows[i]
        name = r[0] if r[0] else cur_station
        cur_station = name  # 총계 행의 두 번째(하차) 행도 올바르게 식별되도록 항상 갱신 (버그 수정 포인트)
        if name is None or name == "1호선 계":
            i += 1
            continue
        gubun = r[1]
        hourly = [int(v) if v is not None else 0 for v in r[3:23]]
        out.setdefault(name, {})[gubun] = hourly
        i += 1
    wb.close()
    return hour_labels, out


def load(pm25_path: str, traffic_path: str, period_label: str = "2026-04"):
    Base.metadata.create_all(bind=engine)

    pm25 = parse_pm25(pm25_path)
    hour_labels, traffic = parse_traffic(traffic_path)

    line_order_names = list(traffic.keys())  # 엑셀 상 등장 순서 = 노선 순서
    assert traffic["송도달빛축제공원"]["하차"] and sum(traffic["송도달빛축제공원"]["하차"]) < 1_000_000, \
        "송도달빛축제공원 하차 합계가 비정상적으로 큽니다 — '1호선 계' 오염 가능성, 파싱 로직을 확인하세요."

    db: Session = SessionLocal()
    try:
        db.query(models.PM25Hourly).delete()
        db.query(models.TrafficHourly).delete()
        db.query(models.Station).delete()
        db.commit()

        station_objs: dict[str, models.Station] = {}
        for idx, name in enumerate(line_order_names):
            lat, lon = LINE1_COORDS.get(name, (0.0, 0.0))
            st = models.Station(
                name=name, line_order=idx, lat=lat, lon=lon, has_pm25=name in pm25,
            )
            db.add(st)
            station_objs[name] = st
        db.commit()

        pm_count = 0
        for name, records in pm25.items():
            st = station_objs.get(name)
            if st is None:
                continue
            for ts, val in records:
                db.add(models.PM25Hourly(station_id=st.id, ts=ts, value=val))
                pm_count += 1
        db.commit()

        tr_count = 0
        for name, directions in traffic.items():
            st = station_objs.get(name)
            if st is None:
                continue
            for direction, hourly in directions.items():
                for h_idx, (label, cnt) in enumerate(zip(hour_labels, hourly)):
                    db.add(models.TrafficHourly(
                        station_id=st.id, direction=direction, hour_label=label,
                        hour_index=h_idx, count=cnt, period_label=period_label,
                    ))
                    tr_count += 1
        db.commit()

        print(f"적재 완료: 역 {len(station_objs)}개, PM2.5 레코드 {pm_count}건, 통행량 레코드 {tr_count}건")
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="초미세먼지/통행량 엑셀 → PostgreSQL 적재")
    parser.add_argument("--pm25", required=True, help="초미세먼지 월보 xlsx 경로")
    parser.add_argument("--traffic", required=True, help="역·시간대별 통행량 xlsx 경로")
    parser.add_argument("--period", default="2026-04", help="통행량 자료 기준 년월 라벨")
    args = parser.parse_args()
    load(args.pm25, args.traffic, args.period)
